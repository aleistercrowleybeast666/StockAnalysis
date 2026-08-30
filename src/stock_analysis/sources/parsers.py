from __future__ import annotations

import io
import json
import re
import warnings
from html.parser import HTMLParser
from typing import Any

from openpyxl import load_workbook

from stock_analysis.sources.base import SourceSchemaError


def Parser_ParseJsonOrJsonp(payload: str) -> Any:
    text = payload.strip().lstrip("\ufeff")
    if not text:
        raise SourceSchemaError("空 JSON/JSONP 响应")
    if text[0] in "[{":
        body = text
    else:
        match = re.fullmatch(r"[A-Za-z_$][\w.$]*\s*\((.*)\)\s*;?", text, re.DOTALL)
        if not match:
            raise SourceSchemaError("JSONP 包装格式无法识别")
        body = match.group(1)
    try:
        return json.loads(body)
    except json.JSONDecodeError as error:
        raise SourceSchemaError("JSON/JSONP 内容损坏") from error


def Parser_ParseHkexWorkbook(payload: bytes) -> list[dict[str, str]]:
    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as error:
        raise SourceSchemaError("港交所证券列表不是有效 XLSX") from error
    sheet = workbook.active
    try:
        reset_dimensions = getattr(sheet, "reset_dimensions", None)
        if callable(reset_dimensions):
            reset_dimensions()
        header_row: int | None = None
        headers: list[str] = []
        for row_index, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=20, values_only=True), 1
        ):
            normalized = [
                str(value).strip() if value is not None else "" for value in row
            ]
            if "Stock Code" in normalized and "Category" in normalized:
                header_row = row_index
                headers = normalized
                break
        if header_row is None:
            raise SourceSchemaError("港交所证券列表缺少预期表头")
        required = ("Stock Code", "Name of Securities", "Category", "Sub-Category")
        index = {name: headers.index(name) for name in required}
        currency_index = next(
            (
                headers.index(name)
                for name in ("Trading Currency", "Currency")
                if name in headers
            ),
            None,
        )
        candidates: list[dict[str, str]] = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            values = ["" if value is None else str(value).strip() for value in row]
            if len(values) <= max(index.values()):
                continue
            category = values[index["Category"]]
            subcategory = values[index["Sub-Category"]]
            if category != "Equity" or not subcategory.startswith("Equity Securities"):
                continue
            raw_code = values[index["Stock Code"]].split(".")[0]
            code = raw_code.zfill(5)
            if not raw_code or not code.isdigit():
                continue
            item = {
                "code": code,
                "name": values[index["Name of Securities"]],
                "category": category,
                "subcategory": subcategory,
            }
            if currency_index is not None and currency_index < len(values):
                item["currency"] = values[currency_index]
            candidates.append(item)
        result: list[dict[str, str]] = []
        issuer_index: dict[str, int] = {}
        for item in candidates:
            normalized_name = re.sub(r"\s+", " ", item["name"].upper()).strip()
            if normalized_name.endswith(("-SWR", "-WR")):
                issuer_name = normalized_name[:-1]
            elif normalized_name.endswith("-R"):
                issuer_name = normalized_name[:-2]
            else:
                issuer_name = normalized_name
            existing_index = issuer_index.get(issuer_name)
            if existing_index is None:
                issuer_index[issuer_name] = len(result)
                result.append(item)
                continue
            existing = result[existing_index]
            existing_hkd = existing.get("currency", "").upper() in {"HKD", "HK$"}
            candidate_hkd = item.get("currency", "").upper() in {"HKD", "HK$"}
            if candidate_hkd and not existing_hkd:
                result[existing_index] = item
        return result
    finally:
        workbook.close()


def Parser_ParseSzseStockWorkbook(payload: bytes) -> list[dict[str, str]]:
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            workbook = load_workbook(io.BytesIO(payload), data_only=True)
    except Exception as error:
        raise SourceSchemaError("深交所股票列表不是有效 XLSX") from error
    sheet = workbook.active
    header_row: int | None = None
    headers: list[str] = []
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1
    ):
        normalized = [str(value).strip() if value is not None else "" for value in row]
        if "A股代码" in normalized and "A股简称" in normalized:
            header_row = row_index
            headers = normalized
            break
    if header_row is None:
        raise SourceSchemaError("深交所股票列表缺少预期表头")
    required = ("A股代码", "A股简称", "A股上市日期", "所属行业")
    index = {name: headers.index(name) for name in required}
    result: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        values = ["" if value is None else str(value).strip() for value in row]
        if len(values) <= max(index.values()):
            continue
        raw_code = values[index["A股代码"]].split(".")[0]
        code = raw_code.zfill(6)
        name = re.sub(r"\s+", "", values[index["A股简称"]])
        if not raw_code or not code.isdigit() or not name:
            continue
        result.append(
            {
                "code": code,
                "name": name,
                "listing_date": values[index["A股上市日期"]],
                "industry": values[index["所属行业"]],
            }
        )
    if not result:
        raise SourceSchemaError("深交所股票列表没有有效 A 股记录")
    return result


class _TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() == "tr":
            self._row = []
        elif tag.lower() in {"td", "th"} and self._row is not None:
            self._cell = []

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"td", "th"} and self._cell is not None and self._row is not None:
            self._row.append("".join(self._cell).strip())
            self._cell = None
        elif tag.lower() == "tr" and self._row is not None:
            self.rows.append(self._row)
            self._row = None


def Parser_ParseHtmlTable(payload: str) -> list[list[str]]:
    parser = _TableParser()
    parser.feed(payload)
    if not parser.rows:
        raise SourceSchemaError("HTML 中没有可解析表格行")
    return parser.rows
