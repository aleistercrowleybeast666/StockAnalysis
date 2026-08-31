from __future__ import annotations

from copy import copy

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_TITLE_FILL = PatternFill("solid", fgColor="4A9BD8")
_SUBTITLE_FILL = PatternFill("solid", fgColor="EAF4FC")
_NEUTRAL_GROUP_FILL = PatternFill("solid", fgColor="D9E2F3")
_NEUTRAL_FIELD_FILL = PatternFill("solid", fgColor="EDF1F7")
_FINANCIAL_GROUP_FILL = PatternFill("solid", fgColor="A9D18E")
_FINANCIAL_FIELD_FILL = PatternFill("solid", fgColor="E2F0D9")
_MARKET_GROUP_FILL = PatternFill("solid", fgColor="9DC3E6")
_MARKET_FIELD_FILL = PatternFill("solid", fgColor="DDEBF7")
_WHITE_FONT = Font(name="等线", size=11, bold=True, color="FFFFFF")
_GROUP_FONT = Font(name="等线", size=10, bold=True, color="1F2937")
_HEADER_FONT = Font(name="等线", size=10, bold=True, color="1F2937")
_BODY_FONT = Font(name="等线", size=10, color="000000")
_LINK_FONT = Font(name="等线", size=10, color="008000")
_THIN_GRAY = Side(style="thin", color="CAD5E2")

MAIN_GROUPS = (
    ("A3:G3", "公司信息", _NEUTRAL_GROUP_FILL),
    ("H3:J3", "营业收入", _FINANCIAL_GROUP_FILL),
    ("K3:M3", "毛利率", _FINANCIAL_GROUP_FILL),
    ("N3:Q3", "归母净利润", _FINANCIAL_GROUP_FILL),
    ("R3:T3", "经营活动现金流", _FINANCIAL_GROUP_FILL),
    ("U3:AE3", "上市发行与市场交易", _MARKET_GROUP_FILL),
)


def _FieldFill_Get(column: int) -> PatternFill:
    if column <= 7:
        return _NEUTRAL_FIELD_FILL
    if column <= 20:
        return _FINANCIAL_FIELD_FILL
    return _MARKET_FIELD_FILL


def Formatting_MainSheet(sheet: Worksheet, last_row: int) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "H5"
    sheet.auto_filter.ref = f"A4:AE{max(4, last_row)}"
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 42
    sheet.row_dimensions[3].height = 24
    sheet.row_dimensions[4].height = 52
    sheet.merge_cells("A1:AE1")
    sheet.merge_cells("A2:AE2")
    sheet["A1"].fill = _TITLE_FILL
    sheet["A1"].font = Font(name="等线", size=13, bold=True, color="FFFFFF")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet["A2"].fill = _SUBTITLE_FILL
    sheet["A2"].font = Font(name="等线", size=10, color="1F2937")
    sheet["A2"].alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )

    for merged_range, title, fill in MAIN_GROUPS:
        sheet.merge_cells(merged_range)
        start = merged_range.split(":", 1)[0]
        cell = sheet[start]
        cell.value = title
        cell.fill = copy(fill)
        cell.font = copy(_GROUP_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=_THIN_GRAY)

    for column in range(1, 32):
        cell = sheet.cell(4, column)
        cell.fill = copy(_FieldFill_Get(column))
        cell.font = copy(_HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=_THIN_GRAY)

    linked_columns = {
        6,
        8,
        14,
        18,
        21,
        22,
        23,
        24,
        25,
        26,
        28,
        29,
        30,
        31,
    }
    percentage_columns = {9, 10, 11, 15, 16, 17, 19, 20, 27}
    point_columns = {12, 13}
    amount_columns = {8, 14, 18, 21, 25, 29, 30, 31}
    price_columns = {23, 26}
    for row in range(5, last_row + 1):
        sheet.row_dimensions[row].height = 21
        for column in range(1, 32):
            cell = sheet.cell(row, column)
            cell.font = copy(_LINK_FONT if column in linked_columns else _BODY_FONT)
            cell.alignment = Alignment(
                horizontal="left" if column in {1, 2, 3, 4, 5, 7} else "right",
                vertical="center",
            )
            if column in percentage_columns:
                cell.number_format = "0.00%;-0.00%;0.00%"
            elif column in point_columns:
                cell.number_format = "0.00;-0.00;0.00"
            elif column in amount_columns:
                cell.number_format = "#,##0.00;-#,##0.00;0.00"
            elif column in price_columns or column == 24:
                cell.number_format = "#,##0.000;-#,##0.000;0.000"
            elif column == 28:
                cell.number_format = "#,##0"
            elif column in {6, 22}:
                cell.number_format = "yyyy-mm-dd"
            elif column == 1:
                cell.number_format = "@"

    widths = {
        "A": 12,
        "B": 18,
        "C": 12,
        "D": 20,
        "E": 38,
        "F": 13,
        "G": 11,
        "H": 16,
        "I": 14,
        "J": 18,
        "K": 11,
        "L": 20,
        "M": 20,
        "N": 16,
        "O": 13,
        "P": 16,
        "Q": 19,
        "R": 19,
        "S": 19,
        "T": 22,
        "U": 15,
        "V": 13,
        "W": 11,
        "X": 13,
        "Y": 16,
        "Z": 13,
        "AA": 13,
        "AB": 19,
        "AC": 19,
        "AD": 18,
        "AE": 22,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:4"


def Formatting_AuxiliarySheet(sheet: Worksheet, last_column: int) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    for column in range(1, last_column + 1):
        cell = sheet.cell(1, column)
        cell.fill = copy(_TITLE_FILL)
        cell.font = copy(_WHITE_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(column)].width = 18
    sheet.row_dimensions[1].height = 34


def Formatting_AddHeaderComments(sheet: Worksheet) -> None:
    ipo_comment = (
        "老上市、介绍上市、重组上市、重新上市或二次上市公司可能没有可比的标准 IPO "
        "发行资料。主数据源及备用公开来源均无法可靠确认时保持空白，不使用当前股本或"
        "其他不等价字段替代。"
    )
    gross_margin_comment = (
        "“-”表示该指标不适用，例如银行、保险、证券等金融企业，或营业收入为非正数而"
        "普通毛利率缺乏可比意义；空白表示指标理论上适用但公开数据源未取得。"
    )
    three_year_comment = (
        "“-”表示所需历史值已取得但 CAGR 基期/本期为非正数，或已确认公司本身不足三个"
        "完整年度，指标在数学上不适用；空白表示理论上应可计算，但本次主备公开来源均未"
        "取得所需历史数据。"
    )
    comments = {
        "板块": (
            "A股按交易所与证券代码规则归类；港股仅使用 HKEX 官方名单中的主板/GEM 分类，"
            "不根据代码、规模或名称猜测。"
        ),
        "行业": "优先使用证券池批量元数据；缺失时才按市场使用公司资料备源。",
        "概念": (
            "主表最多显示 12 个去重标签，完整列表见隐藏工作表“概念明细”。"
            "港股为公开页面的主题/相关指数标签，不等同于 A 股概念分类。"
        ),
        "营业收入同比": "上期营业收入大于 0 时：本期/上期-1。",
        "营业收入三年 CAGR": (
            "本期与三年前营业收入均大于 0 时：(本期/三年前)^(1/3)-1；"
            f"{three_year_comment}"
        ),
        "毛利率": gross_margin_comment,
        "毛利率同比变化（百分点）": (
            f"{gross_margin_comment}适用时为本期毛利率减上期毛利率，以百分点显示。"
        ),
        "毛利率三年变化（百分点）": (
            f"{gross_margin_comment}适用时为本期毛利率减三年前毛利率，不计算 CAGR；"
            f"{three_year_comment}"
        ),
        "归母净利润同比": "上期归母净利润不为 0 时：(本期-上期)/ABS(上期)。",
        "归母净利润三年 CAGR": three_year_comment,
        "经营活动现金流三年 CAGR": three_year_comment,
        "上市日期": (
            "空白表示主备源均未取得。港股发行时间过早时，公开网站可能没有历史记录。"
        ),
        "发行价": ipo_comment,
        "发行股数": ipo_comment,
        "发行时总市值": (
            f"{ipo_comment}有可靠资料时仅按发行价×明确的发行后总股本计算。"
        ),
        "市值增长率": (
            f"{ipo_comment}有可靠发行时总市值时按最新总市值/发行时总市值-1。"
        ),
    }
    if sheet.title == "港股":
        block_trade_comment = (
            "港股大宗交易采用公开网页可验证的 Block Trades/大额成交历史数据累计。由于"
            "免费公开历史记录未必覆盖完整年度，只有能够确认所选年度统计完整性时才填写"
            "数值；确认完整年度无记录时填写 0；无法证明完整覆盖全年时保持空白。空白不"
            "代表没有交易，也不等同于数值 0；详细来源与口径见第三页“数据来源说明”。"
        )
        flow_comment = (
            "近五日按最近 5 个有效交易日累计，近一月按最近 20 个有效交易日累计；港股"
            "资金流采用公开数据平台口径，与 A 股主力资金口径不保证完全一致，空白表示"
            "主备来源均未取得足够历史数据。"
        )
    else:
        block_trade_comment = (
            "按所选年度完整汇总公开大宗交易记录；确认完整年度无记录时填写 0，主备源"
            "无法验证完整年度时保持空白，空白不等同于数值 0。"
        )
        flow_comment = (
            "近五日按最近 5 个有效交易日累计，近一月按最近 22 个有效交易日累计；空白"
            "表示公开主备资金流来源均未取得足够历史数据。"
        )
    comments["当年累计大宗交易笔数"] = block_trade_comment
    comments["当年累计大宗交易金额"] = block_trade_comment
    comments["近五个交易日资金净额"] = flow_comment
    one_month_header = (
        "近一月资金净额（最近20个交易日）"
        if sheet.title == "港股"
        else "近一月资金净额（最近22个交易日）"
    )
    comments[one_month_header] = flow_comment
    for cell in sheet[4]:
        text = comments.get(str(cell.value))
        if text:
            cell.comment = Comment(text, "StockAnalysis")
