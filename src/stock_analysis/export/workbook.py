from __future__ import annotations

import json
import platform
import re
import sys
from collections import Counter
from copy import copy
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from stock_analysis.config.models import AppConfig
from stock_analysis.domain.enums import (
    DataStatus,
    Market,
    TableSortMode,
    WorkbookExportResult,
)
from stock_analysis.domain.fields import (
    FLOW_FIVE_DAY_FIELD,
    FLOW_ONE_MONTH_A_FIELD,
    FLOW_ONE_MONTH_HK_FIELD,
    FlowOneMonthField_Get,
)
from stock_analysis.domain.models import AnalysisRecord
from stock_analysis.export.formatting import (
    Formatting_AddHeaderComments,
    Formatting_AuxiliarySheet,
    Formatting_MainSheet,
)
from stock_analysis.export.resources import Template_GetPath
from stock_analysis.version import APP_DISPLAY_NAME, APP_INTERNAL_NAME, __version__

MAIN_HEADERS = [
    "证券代码",
    "公司名称",
    "实际报告期",
    "报表币种",
    "营业收入",
    "营业收入同比",
    "营业收入三年 CAGR",
    "毛利率",
    "毛利率同比变化（百分点）",
    "毛利率三年变化（百分点）",
    "归母净利润",
    "归母净利率",
    "归母净利润同比",
    "归母净利润三年 CAGR",
    "经营活动现金流净额",
    "经营活动现金流同比",
    "经营活动现金流三年 CAGR",
    "最新总市值",
    "上市日期",
    "发行价",
    "发行股数",
    "发行时总市值",
    "最新可得价格",
    "市值增长率",
    "当年累计大宗交易笔数",
    "当年累计大宗交易金额",
    FLOW_FIVE_DAY_FIELD,
    FLOW_ONE_MONTH_A_FIELD,
]

_COVERAGE_COMMON_FIELDS = tuple(MAIN_HEADERS[4:-1])

COVERAGE_FIELDS_BY_MARKET = {
    Market.A_SHARE: (*_COVERAGE_COMMON_FIELDS, FLOW_ONE_MONTH_A_FIELD),
    Market.HK: (*_COVERAGE_COMMON_FIELDS, FLOW_ONE_MONTH_HK_FIELD),
}

_FINANCIAL_COVERAGE_FIELDS = set(MAIN_HEADERS[4:17])
_DERIVED_COVERAGE_FIELDS = {
    "营业收入同比",
    "营业收入三年 CAGR",
    "毛利率",
    "毛利率同比变化（百分点）",
    "毛利率三年变化（百分点）",
    "归母净利率",
    "归母净利润同比",
    "归母净利润三年 CAGR",
    "经营活动现金流同比",
    "经营活动现金流三年 CAGR",
    "市值增长率",
}
_COVERAGE_FIELD_GROUPS = {
    **{field: {"年度财务"} for field in _FINANCIAL_COVERAGE_FIELDS},
    "最新总市值": {"最新行情"},
    "上市日期": {"上市与发行信息"},
    "发行价": {"上市与发行信息"},
    "发行股数": {"上市与发行信息"},
    "发行时总市值": {"上市与发行信息"},
    "最新可得价格": {"最新行情"},
    "市值增长率": {"上市与发行信息", "最新行情"},
    "当年累计大宗交易笔数": {"大宗交易"},
    "当年累计大宗交易金额": {"大宗交易"},
    FLOW_FIVE_DAY_FIELD: {"资金流"},
    FLOW_ONE_MONTH_A_FIELD: {"资金流"},
    FLOW_ONE_MONTH_HK_FIELD: {"资金流"},
}

_COVERAGE_THRESHOLDS = {
    Market.A_SHARE: {
        **{field: 0.50 for field in COVERAGE_FIELDS_BY_MARKET[Market.A_SHARE]},
        "发行时总市值": 0.85,
        "市值增长率": 0.85,
        "当年累计大宗交易笔数": 1.0,
        "当年累计大宗交易金额": 1.0,
        FLOW_FIVE_DAY_FIELD: 0.90,
        FLOW_ONE_MONTH_A_FIELD: 0.90,
    },
    Market.HK: {
        field: 0.50 for field in COVERAGE_FIELDS_BY_MARKET[Market.HK]
    },
}


def MainHeaders_Get(market: Market, financial_year: int | None = None) -> list[str]:
    headers = list(MAIN_HEADERS)
    headers[-1] = FlowOneMonthField_Get(market)
    if financial_year is not None:
        headers[4] = f"{financial_year}年营业收入"
    return headers

HISTORY_HEADERS = [
    "市场",
    "证券代码",
    "公司名",
    "本期报告日",
    "标准币种",
    "原币种",
    "本期营业收入",
    "上期营业收入",
    "三年前营业收入",
    "本期营业成本",
    "上期营业成本",
    "三年前营业成本",
    "本期归母净利润",
    "上期归母净利润",
    "三年前归母净利润",
    "本期经营现金流",
    "上期经营现金流",
    "三年前经营现金流",
    "最新市值",
    "上市日期",
    "发行价",
    "发行股数",
    "发行时市值",
    "现价",
    "大宗交易数量",
    "大宗交易金额",
    "近五日资金净额",
    "近一月资金净额",
    "是否合并口径",
    "是否追溯调整",
    "财务质量说明",
    "行情日期",
]


def _CellValue(period: object | None, attribute: str) -> object | None:
    return getattr(period, attribute, None) if period is not None else None


def _ExcelDatetime(value: datetime) -> datetime:
    return value.astimezone(UTC).replace(tzinfo=None) if value.tzinfo else value


def _Workbook_Prepare():
    workbook = load_workbook(Template_GetPath())
    first = workbook.active
    first.title = "A股"
    for table_name in list(first.tables):
        del first.tables[table_name]
    for row in first.iter_rows():
        for cell in row:
            cell.value = None
            cell.comment = None
    for merged_range in list(first.merged_cells.ranges):
        first.unmerge_cells(str(merged_range))
    second = workbook.copy_worksheet(first)
    second.title = "港股"
    for name in (
        "数据来源说明",
        "历史数据",
        "数据来源",
        "异常记录",
        "运行信息",
    ):
        if name in workbook.sheetnames:
            del workbook[name]
        workbook.create_sheet(name)
    return workbook


def _StatusValue_Get(
    record: AnalysisRecord,
    field_name: str,
    value: object | None,
    *,
    divisor: float | None = None,
    multiplier: float | None = None,
) -> object | None:
    status = record.field_statuses.get(field_name)
    if status is DataStatus.NOT_APPLICABLE:
        return "-"
    if value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        if divisor:
            number /= divisor
        if multiplier:
            number *= multiplier
        return number
    return value


def _MainRow_Values(record: AnalysisRecord) -> list[object | None]:
    current = record.current
    quote = record.quote
    ipo = record.ipo
    block = record.block_trade
    flow = record.flow
    metrics = record.metrics
    one_month_field = FlowOneMonthField_Get(record.security.market)
    listing_date = (ipo.listing_date if ipo else None) or record.security.listing_date
    return [
        _CellValue(current, "report_end"),
        _CellValue(current, "currency"),
        _StatusValue_Get(record, "营业收入", _CellValue(current, "revenue"), divisor=1e8),
        _StatusValue_Get(record, "营业收入同比", metrics.revenue_growth),
        _StatusValue_Get(record, "营业收入三年 CAGR", metrics.revenue_cagr),
        _StatusValue_Get(record, "毛利率", metrics.gross_margin),
        _StatusValue_Get(
            record,
            "毛利率同比变化（百分点）",
            metrics.gross_margin_yoy_change,
            multiplier=100,
        ),
        _StatusValue_Get(
            record,
            "毛利率三年变化（百分点）",
            metrics.gross_margin_three_year_change,
            multiplier=100,
        ),
        _StatusValue_Get(
            record, "归母净利润", _CellValue(current, "parent_net_profit"), divisor=1e8
        ),
        _StatusValue_Get(record, "归母净利率", metrics.net_margin),
        _StatusValue_Get(record, "归母净利润同比", metrics.profit_growth),
        _StatusValue_Get(record, "归母净利润三年 CAGR", metrics.profit_cagr),
        _StatusValue_Get(
            record,
            "经营活动现金流净额",
            _CellValue(current, "operating_cash_flow"),
            divisor=1e8,
        ),
        _StatusValue_Get(record, "经营活动现金流同比", metrics.cash_growth),
        _StatusValue_Get(record, "经营活动现金流三年 CAGR", metrics.cash_cagr),
        _StatusValue_Get(record, "最新总市值", _CellValue(quote, "market_cap"), divisor=1e8),
        _StatusValue_Get(record, "上市日期", listing_date),
        _StatusValue_Get(record, "发行价", _CellValue(ipo, "issue_price")),
        _StatusValue_Get(record, "发行股数", _CellValue(ipo, "issued_shares"), divisor=1e8),
        _StatusValue_Get(record, "发行时总市值", _CellValue(ipo, "issue_market_cap"), divisor=1e8),
        _StatusValue_Get(record, "最新价", _CellValue(quote, "price")),
        _StatusValue_Get(record, "市值增长率", metrics.market_cap_growth),
        _StatusValue_Get(record, "当年累计大宗交易笔数", _CellValue(block, "trade_count")),
        _StatusValue_Get(
            record, "当年累计大宗交易金额", _CellValue(block, "total_amount"), divisor=1e8
        ),
        _StatusValue_Get(
            record, FLOW_FIVE_DAY_FIELD, _CellValue(flow, "five_day_net"), divisor=1e8
        ),
        _StatusValue_Get(
            record,
            one_month_field,
            _CellValue(flow, "one_month_net"),
            divisor=1e8,
        ),
    ]


def _History_Write(sheet, records: list[AnalysisRecord]) -> dict[str, int]:
    sheet.append(HISTORY_HEADERS)
    row_map: dict[str, int] = {}
    for row_index, record in enumerate(records, 2):
        row_map[record.security.key] = row_index
        current = record.current
        previous = record.previous
        base = record.three_year_base
        quote = record.quote
        ipo = record.ipo
        block = record.block_trade
        flow = record.flow
        sheet.append(
            [
                record.security.market.value,
                record.security.code,
                record.security.name,
                _CellValue(current, "report_end"),
                _CellValue(current, "currency"),
                _CellValue(current, "original_currency"),
                _CellValue(current, "revenue"),
                _CellValue(previous, "revenue"),
                _CellValue(base, "revenue"),
                None
                if record.security.is_financial
                else _CellValue(current, "operating_cost"),
                None
                if record.security.is_financial
                else _CellValue(previous, "operating_cost"),
                None
                if record.security.is_financial
                else _CellValue(base, "operating_cost"),
                _CellValue(current, "parent_net_profit"),
                _CellValue(previous, "parent_net_profit"),
                _CellValue(base, "parent_net_profit"),
                _CellValue(current, "operating_cash_flow"),
                _CellValue(previous, "operating_cash_flow"),
                _CellValue(base, "operating_cash_flow"),
                _CellValue(quote, "market_cap"),
                _CellValue(ipo, "listing_date") or record.security.listing_date,
                _CellValue(ipo, "issue_price"),
                _CellValue(ipo, "issued_shares"),
                _CellValue(ipo, "issue_market_cap"),
                _CellValue(quote, "price"),
                _CellValue(block, "trade_count"),
                _CellValue(block, "total_amount"),
                _CellValue(flow, "five_day_net"),
                _CellValue(flow, "one_month_net"),
                _CellValue(current, "is_consolidated"),
                _CellValue(current, "is_restatement"),
                _CellValue(current, "quality_note"),
                _CellValue(quote, "quote_date"),
            ]
        )
    Formatting_AuxiliarySheet(sheet, len(HISTORY_HEADERS))
    return row_map


def _Main_Write(
    sheet,
    market: Market,
    records: list[AnalysisRecord],
    row_map: dict[str, int],
    config: AppConfig,
) -> None:
    market_records = [record for record in records if record.security.market is market]
    market_records = [record for record in market_records if not record.excluded_reason]
    if config.table_sort_mode is TableSortMode.REVENUE:
        market_records.sort(
            key=lambda record: (
                record.current is None or record.current.revenue is None,
                -(record.current.revenue or 0) if record.current else 0,
                record.security.code,
            )
        )
    else:
        market_records.sort(
            key=lambda record: (
                record.quote is None or record.quote.market_cap is None,
                -(record.quote.market_cap or 0) if record.quote else 0,
                record.security.code,
            )
        )
    quote_dates = sorted(
        {record.quote.quote_date for record in market_records if record.quote is not None}
    )
    if not quote_dates:
        quote_date_text = "未取得"
    elif quote_dates[0] == quote_dates[-1]:
        quote_date_text = quote_dates[0].isoformat()
    else:
        quote_date_text = f"{quote_dates[0].isoformat()} 至 {quote_dates[-1].isoformat()}"
    sheet["A1"] = (
        f"{APP_DISPLAY_NAME}｜财务年度：{config.financial_year}｜真实行情日期："
        f"{quote_date_text}｜金额单位：亿元｜股数单位：亿股"
    )
    if market is Market.A_SHARE:
        scope = "范围：上交所、深交所、北交所普通 A 股；不含 B 股、基金、债券等非普通股票。"
    else:
        scope = (
            "范围：港交所普通股；双柜台证券按发行人去重。"
            "港股大宗交易和资金流会自动尝试备源，无法验证时留空。"
        )
    scope += f" 排序：{config.table_sort_mode.value}；“-”表示不适用，空白表示未取得。"
    sheet["A2"] = scope
    headers = MainHeaders_Get(market, config.financial_year)
    for column, header in enumerate(headers, 1):
        sheet.cell(4, column, header)
    for main_row, record in enumerate(market_records, 5):
        sheet.cell(main_row, 1, record.security.code)
        sheet.cell(main_row, 2, record.security.name)
        values = _MainRow_Values(record)
        for column, value in enumerate(values, 3):
            sheet.cell(main_row, column, value)
    last_row = 4 + len(market_records)
    Formatting_MainSheet(sheet, last_row)
    Formatting_AddHeaderComments(sheet)
    sheet.sheet_properties.tabColor = "70AD47" if market is Market.A_SHARE else "5B9BD5"


def _Coverage_ProvenanceFind(record: AnalysisRecord, field_name: str):
    groups = _COVERAGE_FIELD_GROUPS[field_name]
    matches = [
        item
        for item in record.provenance
        if item.field_group in groups or field_name in item.field_statuses
    ]
    if field_name in _DERIVED_COVERAGE_FIELDS:
        return None
    for item in reversed(matches):
        field_status = item.field_statuses.get(field_name, item.status)
        if field_status is DataStatus.OK:
            return item
    return matches[-1] if matches else None


def _Coverage_BlankReasonGet(record: AnalysisRecord, field_name: str) -> str:
    matching_issues = [
        issue.reason
        for issue in record.issues
        if issue.field_name == field_name and issue.reason
    ]
    if matching_issues:
        return matching_issues[-1]
    groups = _COVERAGE_FIELD_GROUPS[field_name]
    matching_provenance = [
        item
        for item in record.provenance
        if item.field_group in groups or field_name in item.field_statuses
    ]
    for item in reversed(matching_provenance):
        field_status = item.field_statuses.get(field_name, item.status)
        if field_status is not DataStatus.OK and item.missing_reason:
            return item.missing_reason
    status = record.field_statuses.get(field_name, DataStatus.MISSING)
    return f"未取得可验证值（字段状态：{status.value}）"


def _Coverage_BlankReasonNormalize(reason: str) -> str:
    normalized = re.sub(r"\s+", " ", str(reason)).strip()
    if "ETNet" in normalized and "无法证明年度区间完整" in normalized:
        return "ETNet 公开列表未覆盖所选年度完整区间，年度值保持空白"
    if "AASTOCKS" in normalized and "页面未提供" in normalized and "完整区间" in normalized:
        return "AASTOCKS 仅提供当日成交信息，无法验证目标年度完整区间"
    if "HKEX Daily Quotations" in normalized and "滚动公开档案" in normalized:
        return "HKEX 官方滚动日报未覆盖该证券所需的完整年度区间，年度值保持空白"
    if "20 日请求失败" in normalized:
        return "港股 20 日公开资金流源请求失败，字段保持空白"
    if "腾讯资金流仅解析到" in normalized:
        return "腾讯旧资金流端点未返回有效历史，已记录为不可用"
    if len(normalized) > 160:
        return f"{normalized[:157]}..."
    return normalized


def _MarketLatestDateText_Get(
    records: list[AnalysisRecord], market: Market, field: str
) -> str:
    dates = []
    for record in records:
        if record.security.market is not market or record.excluded_reason:
            continue
        value = getattr(record, field)
        if value is None:
            continue
        actual_date = value.quote_date if field == "quote" else value.end_date
        dates.append(actual_date)
    return max(dates).isoformat() if dates else "未取得"


def _Coverage_SourceCellsGet(source_counts: Counter[str]) -> list[object | None]:
    ranked = sorted(source_counts.items(), key=lambda item: (-item[1], item[0]))[:3]
    cells: list[object | None] = []
    for index in range(3):
        if index < len(ranked):
            source_name, count = ranked[index]
            cells.extend([source_name, count])
        else:
            cells.extend([None, 0])
    return cells


def _CoverageRows_Build(records: list[AnalysisRecord]) -> list[list[object | None]]:
    rows: list[list[object | None]] = []
    for market in (Market.A_SHARE, Market.HK):
        market_records = [
            record
            for record in records
            if record.security.market is market and not record.excluded_reason
        ]
        market_headers = MainHeaders_Get(market)
        for field_name in COVERAGE_FIELDS_BY_MARKET[market]:
            field_index = market_headers.index(field_name) - 2
            numeric_count = zero_count = dash_count = blank_count = other_count = 0
            source_counts: Counter[str] = Counter()
            blank_reasons: Counter[str] = Counter()
            for record in market_records:
                value = _MainRow_Values(record)[field_index]
                if value == "-":
                    dash_count += 1
                    continue
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    numeric_count += 1
                    if float(value) == 0.0:
                        zero_count += 1
                    if field_name in _DERIVED_COVERAGE_FIELDS:
                        source_counts["派生计算"] += 1
                    else:
                        provenance = _Coverage_ProvenanceFind(record, field_name)
                        source_counts[
                            provenance.source_name if provenance else "来源记录缺失"
                        ] += 1
                    continue
                if value is not None and str(value).strip():
                    other_count += 1
                    provenance = _Coverage_ProvenanceFind(record, field_name)
                    source_counts[
                        provenance.source_name if provenance else "来源记录缺失"
                    ] += 1
                    continue
                blank_count += 1
                blank_reasons[
                    _Coverage_BlankReasonNormalize(
                        _Coverage_BlankReasonGet(record, field_name)
                    )
                ] += 1
            company_count = len(market_records)
            applicable_count = max(0, company_count - dash_count)
            coverage_rate = (
                (numeric_count + other_count) / applicable_count
                if applicable_count
                else 0.0
            )
            reason_text = "；".join(
                f"{reason}（{count}）"
                for reason, count in blank_reasons.most_common(2)
            )
            rows.append(
                [
                    market.value,
                    field_name,
                    company_count,
                    numeric_count,
                    dash_count,
                    blank_count,
                    zero_count,
                    other_count,
                    coverage_rate,
                    *_Coverage_SourceCellsGet(source_counts),
                    reason_text or None,
                ]
            )
    return rows


def _SourceGuide_Write(
    sheet, config: AppConfig, records: list[AnalysisRecord]
) -> None:
    sheet.append([f"{APP_DISPLAY_NAME}｜数据来源与口径说明（版本 {__version__}）"])
    sheet.append(
        [
            "显示约定：“-”仅表示指标对该公司不适用；空白表示已尝试来源但未取得或无法验证；"
            "数值 0 表示来源成功且确认结果为零。程序不读取或写入跨运行缓存。"
            "数据仅供个人分析参考。"
        ]
    )
    sheet.append(
        [
            f"分析年度：{config.financial_year}；交易统计年度：{config.trading_year}；"
            f"表格排序：{config.table_sort_mode.value}；"
            f"行情日期：A股 {_MarketLatestDateText_Get(records, Market.A_SHARE, 'quote')}，"
            f"港股 {_MarketLatestDateText_Get(records, Market.HK, 'quote')}；"
            f"资金流截止：A股 {_MarketLatestDateText_Get(records, Market.A_SHARE, 'flow')}，"
            f"港股 {_MarketLatestDateText_Get(records, Market.HK, 'flow')}。"
        ]
    )
    sheet.append([None])
    coverage_header_row = sheet.max_row + 1
    sheet.append(
        [
            "市场",
            "字段",
            "总公司数",
            "数值数",
            "- 数",
            "空白数",
            "0 数",
            "非数值已取得",
            "覆盖率",
            "主成功来源",
            "主源成功数",
            "第一备用成功来源",
            "第一备源成功数",
            "第二备用成功来源",
            "第二备源成功数",
            "主要空白原因（次数）",
        ]
    )
    coverage_rows = _CoverageRows_Build(records)
    for row in coverage_rows:
        sheet.append(row)
    coverage_last_row = sheet.max_row
    sheet.append([None])
    static_title_row = sheet.max_row + 1
    sheet.append(
        [
            "以下为配置的数据源、回退顺序与计算口径；上表来源列按本次实际成功数排序。"
        ]
    )
    static_header_row = sheet.max_row + 1
    sheet.append(
        [
            "市场",
            "字段/字段组",
            "主数据源",
            "第一备用数据源",
            "第二备用数据源",
            "回退触发条件",
            "计算或统计口径",
            "“-”适用条件",
            "空白的含义",
            "时间口径",
            "备注/公开入口",
        ]
    )
    rows = [
        (
            "A股",
            "证券名单",
            "上交所、深交所、北交所官方列表",
            "东方财富",
            "无可靠第二备源",
            "官方列表请求失败、字段为空或结构不匹配",
            "仅普通 A 股；按上市状态过滤；是否包含 ST 由界面复选框控制",
            "不适用",
            "主备证券池均不可用时任务失败",
            "每次运行重新获取",
            "交易所证券列表 / https://quote.eastmoney.com/",
        ),
        (
            "港股",
            "证券名单",
            "HKEX ListOfSecurities.xlsx",
            "东方财富",
            "无可靠第二备源",
            "HKEX 下载/解析失败、字段为空或结构不匹配",
            "读取工作表真实行；普通股；双柜台优先保留 HKD 柜台",
            "不适用",
            "主备证券池均不可用时任务失败",
            "每次运行重新获取",
            "https://www.hkex.com.hk/Services/Trading/Securities/Securities-Lists",
        ),
        (
            "A股",
            "营业收入、营业成本、毛利率、归母净利润、经营现金流",
            "巨潮资讯（配置官方接口时）",
            "东方财富 F10",
            "未启用未经验证的网页适配器",
            "主源未配置、请求失败、字段为空或年度不匹配",
            "完整年度、合并口径优先；取分析年、上年和三年前数据",
            "金融企业普通毛利率不适用",
            "主备源均未取得；核心营业收入缺失时跳过并补位",
            "所选完整年度、上一年、所选年减三",
            "https://www.cninfo.com.cn/ / https://data.eastmoney.com/",
        ),
        (
            "港股",
            "营业收入、营业成本、毛利率、归母净利润、经营现金流",
            "东方财富港股 F10",
            "A/H 同一发行人映射（仅明确维护映射）",
            "AASTOCKS/港交所年报（候选，未做全市场 PDF 扫描）",
            "主源请求失败、字段为空或年度不匹配；备源不可验证时不猜测",
            "非 HKD 报表按报告日公开汇率换算为 HKD",
            "金融企业普通毛利率不适用",
            "全部可靠来源均未取得；核心营业收入缺失时跳过并补位",
            "公司完整财年；港股自身历史优先，缺年时才用明确同一发行人 A 股合并历史",
            "https://emweb.securities.eastmoney.com/；已验证 03308.HK ↔ 300308.SZ",
        ),
        (
            "A股/港股",
            "同比、三年 CAGR、三年变化",
            "上述年度财务标准化值",
            "无独立外部来源",
            "无",
            "基础年度值完整后由 Python 计算",
            "同比按相邻完整年度；CAGR 按本期和三年前；毛利率差为百分点",
            "历史不足三年或数学上不适用",
            "理论应有基础值但所有来源均未取得",
            "与所选财务年度一致",
            "不依赖 Excel 公式缓存",
        ),
        (
            "A股",
            "最新价格、总市值、行情日期",
            "东方财富批量行情",
            "腾讯行情",
            "无可靠第二备源",
            "主源请求失败、该证券缺失或没有真实时间字段",
            "使用来源返回的真实行情时间 f124/时间字段，不用系统日期冒充",
            "不适用",
            "主备源均未取得；Top N 时无市值者不参与排名",
            "来源实际最近交易日",
            "https://quote.eastmoney.com/ / https://qt.gtimg.cn/",
        ),
        (
            "港股",
            "最新价格、总市值、行情日期",
            "东方财富批量行情",
            "AASTOCKS 网页行情",
            "无可靠第二备源",
            "主源请求失败、该证券缺失或没有真实时间字段",
            "使用网页披露的 Last Updated；金额币种 HKD",
            "不适用",
            "主备源均未取得；Top N 时无市值者不参与排名",
            "来源实际最近交易日；停牌不冒充当天",
            "https://www.aastocks.com/en/stocks/quote/quick-quote.aspx",
        ),
        (
            "A股",
            "上市日期、发行价、发行股数",
            "东方财富 F10",
            "证券列表中的上市日期",
            "未启用未经验证的公司资料页",
            "主源失败或单个字段为空时仅补可验证字段",
            "发行股数保持原字段；人民币双柜台不作为独立 IPO",
            "不适用",
            "所有可靠来源均未取得",
            "历史上市/发行日期",
            "https://emweb.securities.eastmoney.com/",
        ),
        (
            "港股",
            "上市日期、发行价、发行股数",
            "东方财富港股 F10",
            "ETNet 公司资料页 Listing Date/Price",
            "HKEXnews/AASTOCKS（候选，未启用全市场正文扫描）",
            "东方财富上市日期或发行价单字段为空时按字段回退",
            "ETNet 仅补真实披露的上市日期/上市价；不以当前股本代替发行后总股本",
            "不适用",
            "实际尝试的来源均未取得；发行股数不从当前已发行股本推断",
            "历史上市/发行日期",
            "https://www.etnet.com.hk/www/eng/stocks/realtime/quote_ci_brief.php",
        ),
        (
            "A股/港股",
            "发行时总市值、市值增长率",
            "发行资料与最新总市值标准化值",
            "无独立外部来源",
            "无",
            "基础字段完整后由 Python 计算",
            "发行时总市值仅为发行价×明确的发行后总股本；市值增长率据此计算",
            "数学上不适用时",
            "缺发行价、发行后总股本或最新市值；不以募资额/发行股数替代",
            "发行历史值与实际行情日期",
            "计算口径不同于募集资金总额",
        ),
        (
            "A股",
            "当年累计大宗交易笔数、金额",
            "东方财富大宗交易",
            "交易所公开披露（仅在结构可验证时）",
            "无可靠第二备源",
            "主源请求/正文失败或结构不匹配",
            "所选年度完整分页汇总；成功且无记录才写数值 0",
            "不适用",
            "无法完成所选年度可靠汇总",
            "所选交易统计年度",
            "https://data.eastmoney.com/dzjy/",
        ),
        (
            "港股",
            "当年累计大宗/大额交易笔数、金额",
            "ETNet Block Trade 年度分页与逐篇明细",
            "AASTOCKS 当日 Block Trades 页面（仅作证据，不写年度）",
            "HKEX Daily Quotations 官方逐日成交记录",
            "ETNet 不能跨过年度边界时验证 AASTOCKS，再按 HKEX 交易日历和滚动日报回退",
            "ETNet 逐篇汇总；HKEX 口径为 P 单笔≥3000万港元、M/X 单笔≥2000万港元；均须完整覆盖证券当年适用交易日",
            "不适用",
            "任一适用交易日报缺失即留空；完整覆盖且无记录才写 0",
            "所选交易统计年度",
            "https://www.etnet.com.hk/www/eng/stocks/realtime/quote_blocktrade.php / https://www.hkex.com.hk/eng/stat/smstat/dayquot_12m/",
        ),
        (
            "A股",
            "近 5/22 个交易日资金净额",
            "东方财富历史资金流",
            "同花顺资金面诊股公开历史",
            "无可靠第二备源",
            "主源请求失败、字段为空、记录不足或结构不匹配",
            "5 日与 22 日独立判定：有至少 5 日即可写 5 日；只有严格 22 日才写 22 日",
            "不适用",
            "对应窗口主备源均失败或记录不足；5 日成功时不因 22 日缺失而清空",
            "最近 5/22 个有效交易日",
            "https://data.eastmoney.com/zjlx/ / https://doctor.10jqka.com.cn/",
        ),
        (
            "港股",
            "近 5/20 个交易日资金净额",
            "东方财富港股历史资金流（运行级健康探测）",
            "TradeGo 5/20 日公开排行",
            "AASTOCKS Money Flow 5 日页",
            "历史主源断连、记录不足或无法验证连续同口径",
            "三个历史样本全部成功才扩散；5 日与 20 日独立判定；同公司优先使用同一供应商",
            "不适用",
            "无法验证相应窗口；5 日成功时不因 20 日缺失而清空",
            "最近 5/20 个有效交易日",
            "https://quote.eastmoney.com/hk/00700.html / https://data.tradego8.com/indices/MarketMF.aspx / https://www.aastocks.com/en/stocks/analysis/moneyflow.aspx",
        ),
        (
            "港股",
            "人民币双柜台、A/H 发行人映射",
            "HKEX 证券名单与代码/名称标准化",
            "公开发行人信息",
            "无可靠自动第二备源",
            "检测到 -R/-WR/-SWR，或港股历史缺年且命中明确维护的同一法律发行人映射时",
            "双柜台优先保留 HKD 主柜台；A/H 映射仅补港股自身缺失年度，不能模糊匹配名称",
            "不适用",
            "A/H 映射无法可靠确认时不跨代码拼接；真实历史不足时派生三年指标为 -",
            "证券名单与对应财务年度",
            "同一发行人只输出一次，不把人民币柜台当独立 IPO",
        ),
        (
            "港股",
            "非 HKD 财务换算",
            "Frankfurter 历史汇率",
            "无可靠备用源",
            "无",
            "港股财务原币种不是 HKD 时",
            "按报告日公开历史汇率统一换算为 HKD，并保留原币种溯源",
            "不适用",
            "报告日汇率无法取得时相关标准化值留空",
            "实际财务报告日",
            "https://frankfurter.app/",
        ),
        (
            "A股/港股",
            "无缓存、来源审计",
            "本次运行内请求与来源记录",
            "无",
            "无",
            "每次任务开始时重新访问公开来源",
            "只在本次任务内去重完全相同请求；主源成功不请求备源",
            "不适用",
            "空白原因和主备尝试写入隐藏审计页与日志",
            "本次运行",
            "不创建 SQLite、负缓存或原始响应缓存",
        ),
    ]
    for row in rows:
        sheet.append(row)
    Formatting_AuxiliarySheet(sheet, 16)
    for row in (1, 2, 3):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
        sheet.cell(row, 1).alignment = copy(sheet.cell(1, 1).alignment)
    sheet.merge_cells(
        start_row=static_title_row,
        start_column=1,
        end_row=static_title_row,
        end_column=16,
    )
    sheet.cell(static_title_row, 1).font = Font(bold=True, color="1F2937")
    sheet.cell(static_title_row, 1).alignment = Alignment(
        horizontal="left", vertical="center"
    )
    for header_row in (coverage_header_row, static_header_row):
        for column in range(1, 17):
            sheet.cell(header_row, column).fill = copy(sheet.cell(1, 1).fill)
            sheet.cell(header_row, column).font = copy(sheet.cell(1, 1).font)
            sheet.cell(header_row, column).alignment = copy(sheet.cell(1, 1).alignment)
        sheet.row_dimensions[header_row].height = 36
    low_fill = PatternFill("solid", fgColor="FCE8E6")
    all_blank_fill = PatternFill("solid", fgColor="F4CCCC")
    for row in range(coverage_header_row + 1, coverage_last_row + 1):
        market = Market(sheet.cell(row, 1).value)
        field_name = str(sheet.cell(row, 2).value)
        coverage_rate = float(sheet.cell(row, 9).value or 0.0)
        threshold = _COVERAGE_THRESHOLDS[market][field_name]
        company_count = int(sheet.cell(row, 3).value or 0)
        dash_count = int(sheet.cell(row, 5).value or 0)
        blank_count = int(sheet.cell(row, 6).value or 0)
        applicable_count = max(0, company_count - dash_count)
        fill = (
            all_blank_fill
            if applicable_count > 0 and blank_count == applicable_count
            else low_fill
            if coverage_rate < threshold
            else None
        )
        if fill is not None:
            for column in range(1, 17):
                sheet.cell(row, column).fill = copy(fill)
        sheet.cell(row, 9).number_format = "0.00%"
        sheet.cell(row, 16).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        sheet.row_dimensions[row].height = 60
    for row in range(static_header_row + 1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 66
        for column in range(1, 17):
            sheet.cell(row, column).alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )
    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A{coverage_header_row}:P{coverage_last_row}"
    sheet.column_dimensions["A"].width = 13
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["C"].width = 30
    sheet.column_dimensions["D"].width = 32
    sheet.column_dimensions["E"].width = 32
    sheet.column_dimensions["F"].width = 42
    sheet.column_dimensions["G"].width = 54
    sheet.column_dimensions["H"].width = 28
    sheet.column_dimensions["I"].width = 42
    sheet.column_dimensions["J"].width = 28
    sheet.column_dimensions["K"].width = 54
    sheet.column_dimensions["L"].width = 18
    sheet.column_dimensions["M"].width = 30
    sheet.column_dimensions["N"].width = 18
    sheet.column_dimensions["O"].width = 60
    sheet.column_dimensions["P"].width = 60


def _Provenance_Write(sheet, records: list[AnalysisRecord]) -> None:
    headers = [
        "市场",
        "证券代码",
        "公司名",
        "字段组",
        "来源名",
        "来源地址或请求标识",
        "抓取时间",
        "原币种",
        "标准币种",
        "状态",
        "缺失原因",
        "是否近似",
        "主数据源",
        "字段级状态(JSON)",
    ]
    sheet.append(headers)
    for record in records:
        for item in record.provenance:
            sheet.append(
                [
                    item.market.value,
                    item.code,
                    item.company_name,
                    item.field_group,
                    item.source_name,
                    item.source_ref,
                    _ExcelDatetime(item.fetched_at),
                    item.original_currency,
                    item.standard_currency,
                    item.status.value,
                    item.missing_reason,
                    item.approximate,
                    item.primary_source,
                    json.dumps(
                        {
                            field_name: status.value
                            for field_name, status in item.field_statuses.items()
                        },
                        ensure_ascii=False,
                        sort_keys=True,
                    ),
                ]
            )
    Formatting_AuxiliarySheet(sheet, len(headers))
    sheet.column_dimensions["F"].width = 42
    sheet.column_dimensions["K"].width = 42


def _Issues_Write(sheet, records: list[AnalysisRecord]) -> None:
    headers = [
        "市场",
        "证券代码",
        "公司名称",
        "字段名",
        "字段状态",
        "原因",
        "是否核心字段",
        "是否允许为空",
        "主数据源",
        "实际数据源",
        "请求端点",
        "抓取时间",
    ]
    sheet.append(headers)
    for record in records:
        for issue in record.issues:
            sheet.append(
                [
                    issue.market.value,
                    issue.code,
                    issue.company_name,
                    issue.field_name or issue.stage,
                    issue.field_status.value,
                    issue.reason,
                    issue.is_core,
                    issue.optional,
                    issue.primary_source,
                    issue.source_name,
                    issue.endpoint,
                    _ExcelDatetime(issue.fetched_at) if issue.fetched_at else None,
                ]
            )
    Formatting_AuxiliarySheet(sheet, len(headers))
    sheet.column_dimensions["F"].width = 60
    sheet.column_dimensions["K"].width = 42


def _RunInfo_Write(
    sheet,
    config: AppConfig,
    output_path: Path,
    metadata: dict[str, Any],
) -> None:
    sheet.append(["项目", "值"])
    rows = [
        ("程序英文名", APP_INTERNAL_NAME),
        ("程序中文名", APP_DISPLAY_NAME),
        ("程序版本", __version__),
        ("Python 版本", sys.version),
        ("操作系统", platform.platform()),
        ("运行时间", datetime.now(UTC).isoformat()),
        ("配置快照", json.dumps(config.to_dict(), ensure_ascii=False, sort_keys=True)),
        ("成功数", metadata.get("success_count", 0)),
        ("部分缺失数", metadata.get("partial_count", 0)),
        ("失败数", metadata.get("failed_count", 0)),
        ("排除数", metadata.get("excluded_count", 0)),
        ("跨运行缓存", "未使用；每次运行重新访问数据源"),
        ("本次运行内重复请求复用", metadata.get("run_local_reuse", 0)),
        (
            "性能统计",
            json.dumps(metadata.get("performance", {}), ensure_ascii=False, sort_keys=True),
        ),
        ("输出路径", str(output_path)),
    ]
    market_stats = metadata.get("market_stats", {})
    for market in (Market.A_SHARE, Market.HK):
        stats = market_stats.get(market) if isinstance(market_stats, dict) else None
        if stats is None:
            continue
        rows.extend(
            [
                (f"{market.value}识别数", stats.identified_count),
                (f"{market.value}市值可排名数", stats.ranked_count),
                (f"{market.value}候选财务数", stats.candidate_count),
                (f"{market.value}生成数", stats.generated_count),
                (
                    f"{market.value}核心财务跳过数",
                    stats.skipped_no_core_financial_count,
                ),
                (f"{market.value}终止上市过滤数", stats.skipped_delisted_count),
            ]
        )
    for row in rows:
        sheet.append(row)
    Formatting_AuxiliarySheet(sheet, 2)
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 100


def Workbook_Export(
    records: list[AnalysisRecord],
    config: AppConfig,
    output_path: Path,
    metadata: dict[str, Any] | None = None,
) -> WorkbookExportResult:
    output_path = output_path.expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = _Workbook_Prepare()
    history = workbook["历史数据"]
    row_map = _History_Write(history, records)
    _Main_Write(workbook["A股"], Market.A_SHARE, records, row_map, config)
    _Main_Write(workbook["港股"], Market.HK, records, row_map, config)
    _SourceGuide_Write(workbook["数据来源说明"], config, records)
    _Provenance_Write(workbook["数据来源"], records)
    _Issues_Write(workbook["异常记录"], records)
    _RunInfo_Write(workbook["运行信息"], config, output_path, metadata or {})
    for name in ("历史数据", "数据来源", "异常记录", "运行信息"):
        workbook[name].sheet_state = "hidden"
    workbook.active = workbook.sheetnames.index("A股")
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except AttributeError:
        pass
    temporary = output_path.with_name(f"{output_path.stem}.tmp.xlsx")
    workbook.save(temporary)
    temporary.replace(output_path)
    Workbook_Validate(output_path)
    return WorkbookExportResult.SUCCESS


def Workbook_Validate(path: Path) -> None:
    workbook = load_workbook(path, data_only=False, read_only=False)
    expected = {
        "A股",
        "港股",
        "数据来源说明",
        "历史数据",
        "数据来源",
        "异常记录",
        "运行信息",
    }
    if not expected.issubset(set(workbook.sheetnames)):
        raise ValueError("工作簿缺少必要工作表")
    if workbook.sheetnames[:3] != ["A股", "港股", "数据来源说明"]:
        raise ValueError("工作簿前三张工作表顺序不正确")
    if any(workbook[name].sheet_state != "visible" for name in workbook.sheetnames[:3]):
        raise ValueError("工作簿前三张工作表必须可见")
    for sheet_name in ("A股", "港股"):
        sheet = workbook[sheet_name]
        market = Market.A_SHARE if sheet_name == "A股" else Market.HK
        if sheet.max_column != 28:
            raise ValueError(f"{sheet_name} 主表不是 28 列")
        if [sheet.cell(4, column).value for column in range(1, 29)][0] != "证券代码":
            raise ValueError(f"{sheet_name} 表头不正确")
        if sheet["A3"].value != "公司信息" or sheet["E3"].value != "营业收入":
            raise ValueError(f"{sheet_name} 双层分组表头不正确")
        if sheet["AB4"].value != FlowOneMonthField_Get(market):
            raise ValueError(f"{sheet_name} 近一月资金流表头与市场口径不一致")
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "#REF!" in cell.value:
                    raise ValueError(f"{sheet_name}!{cell.coordinate} 包含 #REF!")
    if getattr(workbook, "_external_links", []):
        raise ValueError("工作簿包含外部链接")
