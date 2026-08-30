from __future__ import annotations

from collections.abc import Sequence
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from stock_analysis.config.models import AppConfig
from stock_analysis.domain.enums import DataStatus, Market, MarketScopeMode, PipelineRunResult
from stock_analysis.domain.models import (
    BlockTradeData,
    FinancialPeriod,
    FlowData,
    IPOInfo,
    Quote,
    Security,
)
from stock_analysis.pipeline.runner import PipelineRunner
from stock_analysis.sources.base import (
    MarketDataSource,
    Provenance_Create,
    SourceUnavailableError,
    SourceValue,
)


class RankingSource(MarketDataSource):
    source_name = "ranking-fixture"

    def __init__(self, company_count: int = 5) -> None:
        self.securities = [
            Security(Market.A_SHARE, "SSE", f"{index:06d}", f"公司{index}")
            for index in range(1, company_count + 1)
        ]
        self.market_caps = {
            security.key: float(company_count - index) * 100.0
            for index, security in enumerate(self.securities)
        }
        if company_count == 5:
            self.market_caps = {
                self.securities[0].key: 300.0,
                self.securities[1].key: 400.0,
                self.securities[2].key: None,
                self.securities[3].key: 500.0,
                self.securities[4].key: 200.0,
            }
        self.no_core_keys = {self.securities[3].key} if company_count == 5 else set()
        self.quote_batch_calls = 0
        self.financial_calls: list[str] = []
        self.flow_calls = 0

    def SecurityList_Fetch(self, market: Market, limit: int = 0) -> list[Security]:
        values = self.securities if market is Market.A_SHARE else []
        return values[:limit] if limit > 0 else list(values)

    def Financials_Fetch(
        self, security: Security, years: set[int]
    ) -> SourceValue[list[FinancialPeriod]]:
        self.financial_calls.append(security.key)
        periods: list[FinancialPeriod] = []
        if security.key not in self.no_core_keys:
            revenue = float(int(security.code)) * 1_000_000.0
            periods.append(
                FinancialPeriod(
                    security.key,
                    date(2025, 12, 31),
                    2025,
                    date(2026, 4, 1),
                    "CNY",
                    revenue,
                    revenue * 0.7,
                    revenue * 0.1,
                    revenue * 0.12,
                    original_currency="CNY",
                )
            )
        return SourceValue(
            periods,
            Provenance_Create(
                security,
                "年度财务",
                self.source_name,
                "fixture-financial",
                DataStatus.OK if periods else DataStatus.MISSING,
                standard_currency="CNY",
                missing_reason=None if periods else "没有目标年度",
            ),
        )
    def Quote_Fetch(self, security: Security) -> SourceValue[Quote]:
        market_cap = self.market_caps.get(security.key)
        quote = Quote(security.key, date.today(), 10.0, market_cap, "CNY")
        return SourceValue(
            quote,
            Provenance_Create(
                security,
                "最新行情",
                self.source_name,
                "fixture-quote",
                DataStatus.OK if market_cap is not None else DataStatus.MISSING,
                standard_currency="CNY",
            ),
        )

    def Quotes_Fetch(
        self, securities: Sequence[Security]
    ) -> dict[str, SourceValue[Quote]]:
        self.quote_batch_calls += 1
        return {security.key: self.Quote_Fetch(security) for security in securities}

    def IPO_Fetch(self, security: Security) -> SourceValue[IPOInfo]:
        return SourceValue(
            None,
            Provenance_Create(
                security,
                "上市与发行信息",
                self.source_name,
                "fixture-ipo",
                DataStatus.OPTIONAL_MISSING,
                standard_currency="CNY",
            ),
        )
    def BlockTrade_Fetch(
        self, security: Security, year: int
    ) -> SourceValue[BlockTradeData]:
        value = BlockTradeData(security.key, year, 0, 0.0, "CNY")
        return SourceValue(
            value,
            Provenance_Create(
                security,
                "大宗交易",
                self.source_name,
                "fixture-block",
                DataStatus.OK,
                standard_currency="CNY",
            ),
        )

    def BlockTrades_Fetch(
        self, securities: Sequence[Security], year: int
    ) -> dict[str, SourceValue[BlockTradeData]]:
        return {security.key: self.BlockTrade_Fetch(security, year) for security in securities}

    def Flow_Fetch(self, security: Security) -> SourceValue[FlowData]:
        self.flow_calls += 1
        return SourceValue(
            None,
            Provenance_Create(
                security,
                "资金流",
                self.source_name,
                "fixture-flow",
                DataStatus.OPTIONAL_MISSING,
                standard_currency="CNY",
            ),
        )


class UnavailableFlowSource(RankingSource):
    def Flow_Fetch(self, security: Security) -> SourceValue[FlowData]:
        self.flow_calls += 1
        raise SourceUnavailableError("模拟资金流主备源均不可用")


def _Run(
    tmp_path: Path,
    source: RankingSource,
    scope: MarketScopeMode,
    top_n: int = 2,
):
    summary = PipelineRunner(
        AppConfig(
            financial_year=2025,
            trading_year=2025,
            markets=[Market.A_SHARE],
            a_share_scope_mode=scope,
            a_share_top_n=top_n,
            output_directory=str(tmp_path),
            request_interval=0,
        ),
        source,
    ).run(tmp_path / f"{scope.name}.xlsx")
    return summary


def test_top_n_uses_market_cap_and_backfills_missing_core_financials(
    tmp_path: Path,
) -> None:
    source = RankingSource()
    summary = _Run(tmp_path, source, MarketScopeMode.TOP_MARKET_CAP, 2)
    assert summary.result is PipelineRunResult.SUCCESS
    stats = summary.market_stats[Market.A_SHARE]
    assert stats.generated_count == 2
    assert stats.skipped_no_core_financial_count == 1
    assert stats.candidate_count == 3
    active_codes = [
        record.security.code for record in summary.records if not record.excluded_reason
    ]
    assert active_codes == ["000002", "000001"]
    assert source.quote_batch_calls == 1
    assert source.flow_calls == 2
    assert not any(issue.field_name == "资金流" for issue in summary.issues)

    workbook = load_workbook(summary.output_path, data_only=False)
    assert [workbook["A股"].cell(row, 1).value for row in (5, 6)] == [
        "000002",
        "000001",
    ]
    history = workbook["历史数据"]
    code_to_row = {
        history.cell(row, 2).value: row for row in range(2, history.max_row + 1)
    }
    for code in active_codes:
        row = code_to_row[code]
        assert history.cell(row, 25).value == 0
        assert history.cell(row, 26).value == 0


def test_all_scope_keeps_company_without_market_cap_but_skips_no_core(
    tmp_path: Path,
) -> None:
    source = RankingSource()
    summary = _Run(tmp_path, source, MarketScopeMode.ALL)
    active_codes = {
        record.security.code for record in summary.records if not record.excluded_reason
    }
    assert active_codes == {"000001", "000002", "000003", "000005"}
    stats = summary.market_stats[Market.A_SHARE]
    assert stats.generated_count == 4
    assert stats.ranked_count == 4


def test_all_scope_financial_progress_does_not_reset_between_batches(
    tmp_path: Path,
) -> None:
    source = RankingSource(company_count=450)
    progress = []
    summary = PipelineRunner(
        AppConfig(
            financial_year=2025,
            trading_year=2025,
            markets=[Market.A_SHARE],
            a_share_scope_mode=MarketScopeMode.ALL,
            output_directory=str(tmp_path),
            request_interval=0,
        ),
        source,
        progress.append,
    ).run(tmp_path / "all-progress.xlsx")
    financial_progress = [
        item
        for item in progress
        if item.stage == "年度财务" and item.current_company
    ]
    assert summary.result is PipelineRunResult.SUCCESS
    assert len(financial_progress) == 450
    assert {item.total for item in financial_progress} == {450}
    assert [item.completed for item in financial_progress] == list(range(1, 451))
    planned_progress = [item for item in progress if item.overall_total > 0]
    assert planned_progress
    assert all(
        current.overall_completed / current.overall_total
        <= following.overall_completed / following.overall_total
        for current, following in zip(
            planned_progress, planned_progress[1:], strict=False
        )
    )
    progress_ratios = [
        item.overall_completed / item.overall_total for item in planned_progress
    ]
    assert max(
        following - current
        for current, following in zip(
            progress_ratios, progress_ratios[1:], strict=False
        )
    ) <= 0.1
    quote_completed = next(
        item
        for item in progress
        if item.stage == "获取行情与市值" and item.completed == item.total
    )
    assert quote_completed.overall_completed / quote_completed.overall_total < 0.1
    assert progress[-1].overall_completed == progress[-1].overall_total


def test_top_n_only_fetches_requested_candidates_when_no_backfill_is_needed(
    tmp_path: Path,
) -> None:
    source = RankingSource(company_count=80)
    summary = _Run(tmp_path, source, MarketScopeMode.TOP_MARKET_CAP, 2)
    assert summary.market_stats[Market.A_SHARE].generated_count == 2
    assert len(source.financial_calls) == 2
    assert len(source.financial_calls) < len(source.securities)


def test_flow_source_failures_are_collapsed_into_one_optional_issue(
    tmp_path: Path,
) -> None:
    source = UnavailableFlowSource(company_count=8)
    summary = _Run(
        tmp_path,
        source,
        MarketScopeMode.TOP_MARKET_CAP,
        5,
    )
    flow_issues = [issue for issue in summary.issues if issue.stage == "资金流"]
    assert source.flow_calls == 5
    assert len(flow_issues) == 1
    assert flow_issues[0].optional is True
    assert "共影响 5 家公司" in flow_issues[0].reason
