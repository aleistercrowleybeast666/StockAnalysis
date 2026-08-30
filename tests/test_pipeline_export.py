from __future__ import annotations

import hashlib
from pathlib import Path

from openpyxl import load_workbook

from stock_analysis.common.paths import Resources_GetPath
from stock_analysis.config.models import AppConfig
from stock_analysis.domain.enums import (
    Market,
    MarketScopeMode,
    PipelineRunResult,
    TableSortMode,
)
from stock_analysis.pipeline.runner import PipelineRunner
from stock_analysis.sources.fixture import FixtureSource


def _Sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_fixture_pipeline_creates_valid_workbook(tmp_path: Path) -> None:
    template = Resources_GetPath("templates/分析表.xlsx")
    template_hash = _Sha256(template)
    output = tmp_path / "股票分析表_fixture.xlsx"
    config = AppConfig(
        financial_year=2025,
        trading_year=2025,
        markets=[Market.A_SHARE, Market.HK],
        a_share_scope_mode=MarketScopeMode.ALL,
        hk_scope_mode=MarketScopeMode.ALL,
        output_directory=str(tmp_path),
        fixture_mode=True,
        test_mode=True,
        concurrency=4,
        request_interval=0,
    )
    progress = []
    summary = PipelineRunner(
        config,
        FixtureSource(),
        progress.append,
    ).run(output)

    assert summary.result is PipelineRunResult.SUCCESS
    assert summary.output_path == output
    assert len(summary.records) == 8
    assert output.is_file()
    assert progress[-1].stage == "完成"
    assert _Sha256(template) == template_hash

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == [
        "A股",
        "港股",
        "数据来源说明",
        "历史数据",
        "数据来源",
        "异常记录",
        "运行信息",
    ]
    assert workbook.active.title == "A股"
    for sheet_name in ("历史数据", "数据来源", "异常记录", "运行信息"):
        assert workbook[sheet_name].sheet_state == "hidden"
    for sheet_name in ("A股", "港股"):
        sheet = workbook[sheet_name]
        assert sheet.max_column == 28
        assert sheet.freeze_panes == "E5"
        assert sheet.auto_filter.ref.startswith("A4:AB")
        assert sheet["A3"].value == "公司信息"
        assert sheet["E3"].value == "营业收入"
        assert sheet["R3"].value == "上市发行与市场交易"
        assert sheet["A4"].value == "证券代码"
        assert sheet["V4"].value == "发行时总市值"
        assert sheet["J4"].value == "毛利率三年变化（百分点）"
        assert sheet["E5"].data_type == "n"
        assert isinstance(sheet["E5"].value, (int, float))
        assert sheet["Z5"].number_format == "#,##0.00;-#,##0.00;0.00"
        assert sheet["AA5"].number_format == "#,##0.00;-#,##0.00;0.00"
        assert sheet["E4"].fill.fgColor.rgb.endswith("E2F0D9")
        assert sheet["R4"].fill.fgColor.rgb.endswith("DDEBF7")
        headers = [sheet.cell(4, column).value for column in range(1, 29)]
        assert "字段状态摘要" not in headers
        assert "公司状态" not in headers
        assert "最新可得价格" in headers
        assert headers[-1] == (
            "近一月资金净额（最近22个交易日）"
            if sheet_name == "A股"
            else "近一月资金净额（最近20个交易日）"
        )
        assert sheet["AB4"].comment is not None
        assert ("22" if sheet_name == "A股" else "20") in sheet["AB4"].comment.text
    a_names = [workbook["A股"].cell(row, 2).value for row in range(5, workbook["A股"].max_row + 1)]
    assert a_names[:2] == ["贵州茅台", "宁德时代"]
    assert workbook["A股"]["H8"].value == "-"
    assert workbook["港股"]["V7"].value is None
    assert workbook["港股"]["Y5"].value == 0
    assert workbook["港股"]["Z5"].value == 0
    assert workbook["港股"]["AA5"].value == 0
    assert workbook["异常记录"].max_row == 1
    source_groups = {
        workbook["数据来源"].cell(row, 4).value
        for row in range(2, workbook["数据来源"].max_row + 1)
    }
    assert "资金流" in source_groups
    source_guide = workbook["数据来源说明"]
    assert source_guide.sheet_state == "visible"
    assert source_guide.max_column == 16
    assert source_guide.auto_filter.ref == "A5:P53"
    assert [source_guide.cell(5, column).value for column in range(1, 17)] == [
        "市场",
        "字段",
        "总公司数",
        "数值数",
        "- 数",
        "空白数",
        "0 数",
        "非数值已取得",
        "覆盖率",
        "主成功来源",
        "主源成功数",
        "第一备用成功来源",
        "第一备源成功数",
        "第二备用成功来源",
        "第二备源成功数",
        "主要空白原因（次数）",
    ]
    static_header_row = next(
        row
        for row in range(1, source_guide.max_row + 1)
        if source_guide.cell(row, 1).value == "市场"
        and source_guide.cell(row, 2).value == "字段/字段组"
    )
    assert [source_guide.cell(static_header_row, column).value for column in range(1, 12)] == [
        "市场",
        "字段/字段组",
        "主数据源",
        "第一备用数据源",
        "第二备用数据源",
        "回退触发条件",
        "计算或统计口径",
        "“-”适用条件",
        "空白的含义",
        "时间口径",
        "备注/公开入口",
    ]
    assert source_guide["A6"].value == "A股"
    assert source_guide["B6"].value == "营业收入"
    assert source_guide["I6"].number_format == "0.00%"
    assert "不读取或写入跨运行缓存" in source_guide["A2"].value
    assert "数据仅供个人分析参考" in source_guide["A2"].value
    assert not getattr(workbook, "_external_links", [])


def test_second_fixture_run_starts_without_persistent_cache(tmp_path: Path) -> None:
    config = AppConfig(
        financial_year=2025,
        trading_year=2025,
        markets=[Market.A_SHARE],
        a_share_scope_mode=MarketScopeMode.TOP_MARKET_CAP,
        a_share_top_n=1,
        output_directory=str(tmp_path),
        fixture_mode=True,
        request_interval=0,
    )
    first = PipelineRunner(config, FixtureSource()).run(tmp_path / "first.xlsx")
    second = PipelineRunner(config, FixtureSource()).run(tmp_path / "second.xlsx")
    assert first.output_path is not None
    assert second.output_path is not None
    assert first.performance["run_local_reuse"] == 0
    assert second.performance["run_local_reuse"] == 0
    assert not list(tmp_path.rglob("*.sqlite3"))


def test_final_table_can_sort_by_analysis_year_revenue(tmp_path: Path) -> None:
    output = tmp_path / "revenue-sort.xlsx"
    config = AppConfig(
        financial_year=2025,
        trading_year=2025,
        markets=[Market.A_SHARE],
        a_share_scope_mode=MarketScopeMode.ALL,
        table_sort_mode=TableSortMode.REVENUE,
        output_directory=str(tmp_path),
        fixture_mode=True,
        request_interval=0,
    )
    summary = PipelineRunner(config, FixtureSource()).run(output)
    workbook = load_workbook(output, data_only=True)
    names = [
        workbook["A股"].cell(row, 2).value
        for row in range(5, workbook["A股"].max_row + 1)
    ]
    assert summary.result is PipelineRunResult.SUCCESS
    assert names == ["宁德时代", "平安银行", "贵州茅台", "中芯国际"]
