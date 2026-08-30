from __future__ import annotations

import io
import re
import zipfile

import pytest
from openpyxl import Workbook

from stock_analysis.domain.enums import Market
from stock_analysis.sources.base import SourceSchemaError
from stock_analysis.sources.normalization import (
    Security_ExchangeFromCode,
    Security_FinancialClassify,
    Security_NormalizeCode,
    Security_Secucode,
)
from stock_analysis.sources.parsers import (
    Parser_ParseHkexWorkbook,
    Parser_ParseHtmlTable,
    Parser_ParseJsonOrJsonp,
    Parser_ParseSzseStockWorkbook,
)


def test_json_and_jsonp_parsing() -> None:
    assert Parser_ParseJsonOrJsonp('{"data":{"total":1}}')["data"]["total"] == 1
    assert Parser_ParseJsonOrJsonp('callback_1({"ok":true});')["ok"] is True
    with pytest.raises(SourceSchemaError):
        Parser_ParseJsonOrJsonp("callback(not-json)")


def test_html_table_parsing() -> None:
    rows = Parser_ParseHtmlTable("<table><tr><th>代码</th><th>名称</th></tr><tr><td>1</td><td>A 股</td></tr></table>")
    assert rows == [["代码", "名称"], ["1", "A 股"]]
    with pytest.raises(SourceSchemaError):
        Parser_ParseHtmlTable("<div>no table</div>")


def test_hkex_xlsx_parser_filters_to_ordinary_equity() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["List of Securities"])
    sheet.append([])
    sheet.append(["Stock Code", "Name of Securities", "Category", "Sub-Category"])
    sheet.append([700, "TENCENT", "Equity", "Equity Securities (Main Board)"])
    sheet.append([2800, "TRACKER", "Exchange Traded Products", "Exchange Traded Funds"])
    payload = io.BytesIO()
    workbook.save(payload)
    result = Parser_ParseHkexWorkbook(payload.getvalue())
    assert result == [
        {
            "code": "00700",
            "name": "TENCENT",
            "category": "Equity",
            "subcategory": "Equity Securities (Main Board)",
        }
    ]


def test_hkex_parser_deduplicates_dual_counters_and_prefers_hkd() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Stock Code",
            "Name of Securities",
            "Category",
            "Sub-Category",
            "Trading Currency",
        ]
    )
    sheet.append([80700, "TENCENT-R", "Equity", "Equity Securities (Main Board)", "RMB"])
    sheet.append([700, "TENCENT", "Equity", "Equity Securities (Main Board)", "HKD"])
    sheet.append([89988, "ALIBABA-WR", "Equity", "Equity Securities (Main Board)", "RMB"])
    sheet.append([9988, "ALIBABA-W", "Equity", "Equity Securities (Main Board)", "HKD"])
    payload = io.BytesIO()
    workbook.save(payload)

    result = Parser_ParseHkexWorkbook(payload.getvalue())

    assert [item["code"] for item in result] == ["00700", "09988"]
    assert all(item["currency"] == "HKD" for item in result)


def test_hkex_workbook_parser_ignores_incorrect_declared_dimension() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Stock Code", "Name of Securities", "Category", "Sub-Category"])
    for code in range(1, 301):
        sheet.append(
            [code, f"SAMPLE {code}", "Equity", "Equity Securities (Main Board)"]
        )
    original = io.BytesIO()
    workbook.save(original)
    modified = io.BytesIO()
    with (
        zipfile.ZipFile(original, "r") as source_zip,
        zipfile.ZipFile(modified, "w") as target_zip,
    ):
        for item in source_zip.infolist():
            payload = source_zip.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                payload = re.sub(
                    br'<dimension ref="[^"]+"',
                    b'<dimension ref="A1:D8"',
                    payload,
                    count=1,
                )
            target_zip.writestr(item, payload)
    result = Parser_ParseHkexWorkbook(modified.getvalue())
    assert len(result) == 300
    assert result[-1]["code"] == "00300"


def test_szse_xlsx_parser_maps_a_shares() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["板块", "A股代码", "A股简称", "A股上市日期", "所属行业"])
    sheet.append(["主板", "000001", "平安 银行", "1991-04-03", "J 金融业"])
    sheet.append(["主板", "", "无效行", "", ""])
    payload = io.BytesIO()
    workbook.save(payload)
    result = Parser_ParseSzseStockWorkbook(payload.getvalue())
    assert result == [
        {
            "code": "000001",
            "name": "平安银行",
            "listing_date": "1991-04-03",
            "industry": "J 金融业",
        }
    ]


def test_code_normalization_and_exchange_mapping() -> None:
    assert Security_NormalizeCode("700.HK", Market.HK) == "00700"
    assert Security_NormalizeCode("600519.SH", Market.A_SHARE) == "600519"
    assert Security_ExchangeFromCode("600519") == "SSE"
    assert Security_ExchangeFromCode("300750") == "SZSE"
    assert Security_ExchangeFromCode("832000") == "BSE"
    assert Security_Secucode("700", Market.HK) == "00700.HK"
    assert Security_Secucode("688981", Market.A_SHARE) == "688981.SH"
    with pytest.raises(ValueError):
        Security_NormalizeCode("12", Market.A_SHARE)


def test_financial_classification_uses_industry_then_known_name_patterns() -> None:
    assert Security_FinancialClassify("普通公司", "J 金融业")
    assert Security_FinancialClassify("HSBC HOLDINGS")
    assert Security_FinancialClassify("BOC HONG KONG")
    for code, name in (
        ("02318", "PING AN"),
        ("02628", "CHINA LIFE"),
        ("01299", "AIA"),
        ("00388", "HKEX"),
        ("06030", "CITIC SEC"),
        ("02611", "GTHT"),
        ("06886", "HTSC"),
        ("01776", "GF SEC"),
        ("02328", "PICC P&C"),
        ("02601", "CPIC"),
        ("02378", "PRUDENTIAL"),
        ("00945", "MANULIFE"),
        ("03328", "BANKCOMM"),
        ("02888", "STANCHART"),
        ("00267", "CITIC"),
        ("01339", "PICC GROUP"),
        ("01336", "NCI"),
    ):
        assert Security_FinancialClassify(name, security_code=code)
    assert not Security_FinancialClassify("CKH HOLDINGS")
