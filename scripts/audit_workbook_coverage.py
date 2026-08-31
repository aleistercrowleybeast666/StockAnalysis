#!/usr/bin/env python3
"""Audit coverage of every visible business field in the two market sheets.

The report intentionally distinguishes a true numeric zero from a dash and a
blank cell.  This prevents a failed fetch from being mistaken for a legitimate
zero-value observation.
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from collections.abc import Iterable
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

MARKET_SHEETS = ("A股", "港股")
COMMON_TARGET_FIELDS = (
    "发行时总市值",
    "市值增长率",
    "当年累计大宗交易笔数",
    "当年累计大宗交易金额",
    "近五个交易日资金净额",
)
TARGET_FIELDS_BY_SHEET = {
    "A股": (*COMMON_TARGET_FIELDS, "近一月资金净额（最近22个交易日）"),
    "港股": (*COMMON_TARGET_FIELDS, "近一月资金净额（最近20个交易日）"),
}
CLASSIFICATION_FIELDS = ("板块", "行业", "概念", "最新可得价格")
REQUIRED_FIELDS_BY_SHEET = {
    "A股": (*CLASSIFICATION_FIELDS, *TARGET_FIELDS_BY_SHEET["A股"]),
    "港股": (*CLASSIFICATION_FIELDS, *TARGET_FIELDS_BY_SHEET["港股"]),
}
COVERAGE_THRESHOLDS_BY_SHEET = {
    "A股": {"板块": 0.999, "行业": 0.98, "概念": 0.90, "最新可得价格": 0.99},
    "港股": {"板块": 0.99, "行业": 0.90, "最新可得价格": 0.95},
}
# Kept for callers that build an A-share-compatible fixture.
TARGET_FIELDS = TARGET_FIELDS_BY_SHEET["A股"]
CODE_HEADER = "证券代码"
DASH_VALUES = {"-", "—", "–"}


@dataclass(frozen=True)
class FieldCoverage:
    field: str
    company_count: int
    applicable_count: int
    obtained_count: int
    numeric_count: int
    nonzero_numeric_count: int
    zero_count: int
    dash_count: int
    blank_count: int
    other_text_count: int
    coverage_rate: float
    all_blank: bool
    all_numeric_missing: bool


def _HeaderRow_Find(sheet: Any, search_limit: int = 20) -> int:
    for row_index in range(1, min(sheet.max_row, search_limit) + 1):
        values = {
            str(sheet.cell(row_index, column_index).value).strip()
            for column_index in range(1, sheet.max_column + 1)
            if sheet.cell(row_index, column_index).value is not None
        }
        if CODE_HEADER in values:
            return row_index
    raise ValueError(f"工作表 {sheet.title!r} 未找到表头 {CODE_HEADER!r}")


def _Headers_Index(sheet: Any, header_row: int) -> dict[str, int]:
    result: dict[str, int] = {}
    for column_index in range(1, sheet.max_column + 1):
        value = sheet.cell(header_row, column_index).value
        if value is not None and str(value).strip():
            result[str(value).strip()] = column_index
    return result


def _SheetSnapshot_Read(
    sheet: Any, search_limit: int = 20
) -> tuple[int, dict[str, int], list[tuple[Any, ...]]]:
    """Read a read-only worksheet once; indices in the returned map are zero-based."""

    header_row: int | None = None
    headers: dict[str, int] = {}
    data_rows: list[tuple[Any, ...]] = []
    code_index: int | None = None
    for row_index, values in enumerate(sheet.iter_rows(values_only=True), 1):
        row = tuple(values)
        if header_row is None:
            if row_index > search_limit:
                break
            normalized = [
                str(value).strip() if value is not None else "" for value in row
            ]
            if CODE_HEADER not in normalized:
                continue
            header_row = row_index
            headers = {
                value: index for index, value in enumerate(normalized) if value
            }
            code_index = headers[CODE_HEADER]
            continue
        if code_index is None or code_index >= len(row):
            continue
        code = row[code_index]
        if code is None or not str(code).strip():
            continue
        data_rows.append(row)
    if header_row is None:
        raise ValueError(f"工作表 {sheet.title!r} 未找到表头 {CODE_HEADER!r}")
    return header_row, headers, data_rows


def _DataRows_List(sheet: Any, header_row: int, code_column: int) -> list[int]:
    rows: list[int] = []
    for row_index in range(header_row + 1, sheet.max_row + 1):
        code = sheet.cell(row_index, code_column).value
        if code is None or not str(code).strip():
            continue
        rows.append(row_index)
    return rows


def _Numeric_TryParse(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text or text in DASH_VALUES:
            return None
        if text.endswith("%"):
            text = text[:-1]
        try:
            number = float(text)
        except ValueError:
            return None
        return number if math.isfinite(number) else None
    return None


def _Field_Audit(sheet: Any, rows: Iterable[int], column: int, field: str) -> FieldCoverage:
    row_list = list(rows)
    numeric_count = 0
    nonzero_numeric_count = 0
    zero_count = 0
    dash_count = 0
    blank_count = 0
    other_text_count = 0

    for row_index in row_list:
        value = sheet.cell(row_index, column).value
        if value is None or (isinstance(value, str) and not value.strip()):
            blank_count += 1
            continue
        if isinstance(value, str) and value.strip() in DASH_VALUES:
            dash_count += 1
            continue
        number = _Numeric_TryParse(value)
        if number is None:
            other_text_count += 1
            continue
        numeric_count += 1
        if math.isclose(number, 0.0, abs_tol=1e-12):
            zero_count += 1
        else:
            nonzero_numeric_count += 1

    company_count = len(row_list)
    applicable_count = max(0, company_count - dash_count)
    obtained_count = numeric_count + other_text_count
    coverage_rate = obtained_count / applicable_count if applicable_count else 0.0
    return FieldCoverage(
        field=field,
        company_count=company_count,
        applicable_count=applicable_count,
        obtained_count=obtained_count,
        numeric_count=numeric_count,
        nonzero_numeric_count=nonzero_numeric_count,
        zero_count=zero_count,
        dash_count=dash_count,
        blank_count=blank_count,
        other_text_count=other_text_count,
        coverage_rate=round(coverage_rate, 6),
        all_blank=company_count > 0 and blank_count == company_count,
        all_numeric_missing=applicable_count > 0 and numeric_count == 0,
    )


def _FieldValues_Audit(
    rows: Iterable[tuple[Any, ...]], column: int, field: str
) -> FieldCoverage:
    row_list = list(rows)
    numeric_count = nonzero_numeric_count = zero_count = 0
    dash_count = blank_count = other_text_count = 0
    for row in row_list:
        value = row[column] if column < len(row) else None
        if value is None or (isinstance(value, str) and not value.strip()):
            blank_count += 1
            continue
        if isinstance(value, str) and value.strip() in DASH_VALUES:
            dash_count += 1
            continue
        number = _Numeric_TryParse(value)
        if number is None:
            other_text_count += 1
            continue
        numeric_count += 1
        if math.isclose(number, 0.0, abs_tol=1e-12):
            zero_count += 1
        else:
            nonzero_numeric_count += 1
    company_count = len(row_list)
    applicable_count = max(0, company_count - dash_count)
    obtained_count = numeric_count + other_text_count
    coverage_rate = obtained_count / applicable_count if applicable_count else 0.0
    return FieldCoverage(
        field=field,
        company_count=company_count,
        applicable_count=applicable_count,
        obtained_count=obtained_count,
        numeric_count=numeric_count,
        nonzero_numeric_count=nonzero_numeric_count,
        zero_count=zero_count,
        dash_count=dash_count,
        blank_count=blank_count,
        other_text_count=other_text_count,
        coverage_rate=round(coverage_rate, 6),
        all_blank=company_count > 0 and blank_count == company_count,
        all_numeric_missing=applicable_count > 0 and numeric_count == 0,
    )


def _CellCategory_Get(value: Any) -> str:
    if value is None or (isinstance(value, str) and not value.strip()):
        return "blank"
    if isinstance(value, str) and value.strip() in DASH_VALUES:
        return "dash"
    return "numeric" if _Numeric_TryParse(value) is not None else "other"


def _HkMissingReasons_Build(
    workbook: Any,
    headers: dict[str, int],
    data_rows: list[tuple[Any, ...]],
    fields: tuple[str, ...],
) -> list[dict[str, Any]]:
    code_column = headers[CODE_HEADER]
    name_column = headers.get("公司名称")
    provenance_by_code: dict[str, list[dict[str, str]]] = {}
    if "数据来源" in workbook.sheetnames:
        source_sheet = workbook["数据来源"]
        source_rows = source_sheet.iter_rows(values_only=True)
        header_values = next(source_rows, ())
        source_headers = {
            str(value).strip(): index
            for index, value in enumerate(header_values)
            if value is not None and str(value).strip()
        }
        for values in source_rows:
            market_index = source_headers.get("市场", 0)
            market = str(values[market_index] or "") if market_index < len(values) else ""
            if market != "港股":
                continue
            code_index = source_headers.get("证券代码", 1)
            code = str(values[code_index] or "") if code_index < len(values) else ""
            code = code.zfill(5)
            provenance_by_code.setdefault(code, []).append(
                {
                    key: str(values[column] or "") if column < len(values) else ""
                    for key, column in source_headers.items()
                }
            )
    matrix: list[dict[str, Any]] = []
    for row in data_rows:
        code = str(row[code_column] if code_column < len(row) else "").zfill(5)
        company_name = (
            str(row[name_column] or "")
            if name_column is not None and name_column < len(row)
            else ""
        )
        records = provenance_by_code.get(code, [])
        for field in fields:
            column = headers[field]
            category = _CellCategory_Get(row[column] if column < len(row) else None)
            if category in {"numeric", "other"}:
                continue
            matches = []
            for record in records:
                statuses = record.get("字段级状态(JSON)", "")
                if field in statuses:
                    matches.append(record)
                    continue
                group = record.get("字段组", "")
                flow_match = (
                    field
                    in {"近五个交易日资金净额", "近一月资金净额（最近20个交易日）"}
                    and group == "资金流"
                )
                block_match = (
                    field.startswith("当年累计大宗交易") and group == "大宗交易"
                )
                financial_match = field in {
                    "营业收入同比",
                    "营业收入三年 CAGR",
                    "毛利率",
                    "毛利率同比变化（百分点）",
                    "毛利率三年变化（百分点）",
                    "归母净利润",
                    "归母净利率",
                    "归母净利润同比",
                    "归母净利润三年 CAGR",
                    "经营活动现金流净额",
                    "经营活动现金流同比",
                    "经营活动现金流三年 CAGR",
                } and group == "年度财务"
                financial_match = financial_match or (
                    field.endswith("年营业收入") and group == "年度财务"
                )
                quote_match = field in {"最新总市值", "最新可得价格"} and group == "最新行情"
                classification_match = (
                    field in {"板块", "行业"} and group in {"板块与行业", "行业与概念"}
                ) or (field == "概念" and group in {"概念", "行业与概念"})
                ipo_match = field in {
                    "发行价",
                    "发行股数",
                    "发行时总市值",
                    "市值增长率",
                } and group in {"上市与发行信息", "最新行情"}
                if (
                    flow_match
                    or block_match
                    or ipo_match
                    or financial_match
                    or quote_match
                    or classification_match
                ):
                    matches.append(record)
            last = matches[-1] if matches else {}
            source_names = []
            for item in matches:
                combined_name = item.get("来源名", "")
                source_names.extend(
                    name.strip() for name in combined_name.split("+") if name.strip()
                )
            primary = last.get("主数据源", "") or (source_names[0] if source_names else "")
            fallback_names = [name for name in source_names if name and name != primary]
            reason = last.get("缺失原因", "") or (
                "业务或数学上不适用" if category == "dash" else "未找到字段级来源记录"
            )
            error_text = " ".join(
                [reason, last.get("状态", ""), last.get("来源地址或请求标识", "")]
            )
            program_error = any(
                token in error_text for token in ("错误", "失败", "异常", "解析", "结构")
            )
            public_unavailable = any(
                token in reason
                for token in ("未提供", "未返回", "未取得", "公开", "无法验证", "介绍上市")
            )
            reason_category = (
                "不适用"
                if category == "dash"
                else "程序/解析错误"
                if program_error
                else "公开来源无可验证数据"
                if public_unavailable
                else "来源记录缺失"
                if not matches
                else "来源返回空白/其他"
            )
            matrix.append(
                {
                    "code": code,
                    "company_name": company_name,
                    "field": field,
                    "applicable": category != "dash",
                    "primary_source": primary or None,
                    "attempted_fallbacks": list(dict.fromkeys(fallback_names)),
                    "final_reason": reason,
                    "reason_category": reason_category,
                    "program_error": program_error,
                    "public_source_unavailable": public_unavailable,
                }
            )
    return matrix


def Workbook_Audit(
    path: Path, *, allow_legacy_hk_header: bool = False
) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: dict[str, Any] = {}
        snapshots: dict[str, tuple[dict[str, int], list[tuple[Any, ...]]]] = {}
        for sheet_name in MARKET_SHEETS:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"缺少工作表 {sheet_name!r}")
            sheet = workbook[sheet_name]
            header_row, headers, rows = _SheetSnapshot_Read(sheet)
            numeric_gate_fields = TARGET_FIELDS_BY_SHEET[sheet_name]
            if (
                sheet_name == "港股"
                and allow_legacy_hk_header
                and numeric_gate_fields[-1] not in headers
                and TARGET_FIELDS_BY_SHEET["A股"][-1] in headers
            ):
                numeric_gate_fields = (
                    *COMMON_TARGET_FIELDS,
                    TARGET_FIELDS_BY_SHEET["A股"][-1],
                )
            required_fields = (
                numeric_gate_fields
                if allow_legacy_hk_header
                else REQUIRED_FIELDS_BY_SHEET[sheet_name]
            )
            missing_fields = [field for field in required_fields if field not in headers]
            if missing_fields:
                raise ValueError(
                    f"工作表 {sheet_name!r} 缺少目标字段：{', '.join(missing_fields)}"
                )
            target_fields = tuple(headers)
            coverages = [
                _FieldValues_Audit(rows, headers[field], field)
                for field in target_fields
            ]
            snapshots[sheet_name] = (headers, rows)
            sheets[sheet_name] = {
                "header_row": header_row,
                "company_count": len(rows),
                "numeric_gate_fields": list(numeric_gate_fields),
                "fields": {item.field: asdict(item) for item in coverages},
            }
        empty_numeric_columns = [
            {"market": sheet_name, "field": field_name}
            for sheet_name, sheet_report in sheets.items()
            for field_name in sheet_report["numeric_gate_fields"]
            for field_report in (sheet_report["fields"][field_name],)
            if field_report["all_numeric_missing"]
        ]
        low_coverage_fields = [
            {
                "market": sheet_name,
                "field": field_name,
                "coverage_rate": sheet_report["fields"][field_name]["coverage_rate"],
                "threshold": threshold,
            }
            for sheet_name, thresholds in COVERAGE_THRESHOLDS_BY_SHEET.items()
            for field_name, threshold in thresholds.items()
            for sheet_report in (sheets[sheet_name],)
            if field_name in sheet_report["fields"]
            and sheet_report["fields"][field_name]["applicable_count"] > 0
            and sheet_report["fields"][field_name]["coverage_rate"] < threshold
        ]
        hk_headers, hk_rows = snapshots["港股"]
        hk_fields = tuple(
            field
            for field in sheets["港股"]["fields"]
            if field not in {"证券代码", "公司名称", "实际报告期", "报表币种"}
        )
        return {
            "schema_version": 4,
            "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
            "workbook": str(path.resolve()),
            "coverage_definition": (
                "obtained_count / applicable_count；obtained_count = numeric_count + other_text_count；"
                "applicable_count = company_count - dash_count"
            ),
            "sheets": sheets,
            "hk_missing_reason_matrix": _HkMissingReasons_Build(
                workbook, hk_headers, hk_rows, hk_fields
            ),
            "release_gate": {
                "passed": not empty_numeric_columns and not low_coverage_fields,
                "empty_numeric_columns": empty_numeric_columns,
                "low_coverage_fields": low_coverage_fields,
                "rule": (
                    "有适用公司的目标字段不得整列无数值；"
                    "板块/行业/概念/最新价格须达到市场覆盖率阈值"
                ),
            },
        }
    finally:
        workbook.close()


def _Markdown_Render(report: dict[str, Any]) -> str:
    lines = [
        "# 工作簿全字段覆盖率审计",
        "",
        f"- 文件：`{report['workbook']}`",
        f"- 生成时间：{report['generated_at']}",
        f"- 覆盖率定义：{report['coverage_definition']}",
        f"- 整列数值门禁：{'通过' if report['release_gate']['passed'] else '未通过'}",
        "",
        "数值 0 单独计数；`-` 表示明确不适用；空白表示未取得或未写入，三者不会混算。",
        "",
    ]
    for sheet_name, sheet_report in report["sheets"].items():
        lines.extend(
            [
                f"## {sheet_name}",
                "",
                f"公司数：{sheet_report['company_count']}",
                "",
                "| 字段 | 理论适用 | 已取得 | 数值 | 非零数值 | 零 | `-` | 空白 | 其他文本 | 覆盖率 | 整列空白 | 整列无数值 |",
                "|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|:---:|:---:|",
            ]
        )
        for field, item in sheet_report["fields"].items():
            row_values = dict(item)
            row_values["field"] = field
            row_values["rate"] = item["coverage_rate"]
            row_values["all_blank_label"] = "是" if item["all_blank"] else "否"
            row_values["all_numeric_missing_label"] = (
                "是" if item["all_numeric_missing"] else "否"
            )
            lines.append(
                "| {field} | {applicable_count} | {obtained_count} | {numeric_count} | "
                "{nonzero_numeric_count} | {zero_count} | {dash_count} | "
                "{blank_count} | {other_text_count} | {rate:.2%} | "
                "{all_blank_label} | {all_numeric_missing_label} |".format(**row_values)
            )
        lines.append("")
    if not report["release_gate"]["passed"]:
        lines.extend(["## 未通过门禁的整列", ""])
        for item in report["release_gate"]["empty_numeric_columns"]:
            lines.append(f"- {item['market']}：{item['field']}")
        for item in report["release_gate"].get("low_coverage_fields", []):
            lines.append(
                f"- {item['market']}：{item['field']} 覆盖率 "
                f"{item['coverage_rate']:.2%}，低于 {item['threshold']:.2%}"
            )
        lines.append("")
    matrix = report.get("hk_missing_reason_matrix") or []
    if matrix:
        lines.extend(
            [
                "## 港股逐公司缺失原因矩阵",
                "",
                "| 代码 | 公司 | 字段 | 适用 | 主源 | 已尝试备用源 | 原因分类 | 最终原因 | 程序错误 | 公开来源确无数据 |",
                "|---|---|---|:---:|---|---|---|---|:---:|:---:|",
            ]
        )
        for item in matrix:
            safe = {
                key: str(value or "").replace("|", "\\|").replace("\n", " ")
                for key, value in item.items()
            }
            safe["fallbacks"] = "、".join(item["attempted_fallbacks"]) or "-"
            safe["applicable_label"] = "是" if item["applicable"] else "否"
            safe["program_error_label"] = "是" if item["program_error"] else "否"
            safe["public_label"] = "是" if item["public_source_unavailable"] else "否"
            lines.append(
                "| {code} | {company_name} | {field} | {applicable_label} | "
                "{primary_source} | {fallbacks} | {reason_category} | {final_reason} | "
                "{program_error_label} | {public_label} |".format(**safe)
            )
        lines.append("")
    return "\n".join(lines)


def CoverageReports_Write(
    report: dict[str, Any], json_path: Path, markdown_path: Path
) -> None:
    json_path.parent.mkdir(parents=True, exist_ok=True)
    markdown_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    markdown_path.write_text(_Markdown_Render(report), encoding="utf-8")


def _Arguments_Parse(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="待审计的 .xlsx 文件")
    parser.add_argument("--json", dest="json_path", type=Path, required=True)
    parser.add_argument("--markdown", dest="markdown_path", type=Path, required=True)
    parser.add_argument(
        "--allow-empty-columns",
        action="store_true",
        help="仅用于基线审计；仍写入门禁失败，但命令返回成功",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    arguments = _Arguments_Parse(sys.argv[1:] if argv is None else argv)
    if not arguments.workbook.is_file():
        print(f"文件不存在：{arguments.workbook}", file=sys.stderr)
        return 2
    try:
        report = Workbook_Audit(
            arguments.workbook,
            allow_legacy_hk_header=arguments.allow_empty_columns,
        )
    except (OSError, ValueError) as exc:
        print(f"审计失败：{exc}", file=sys.stderr)
        return 1

    CoverageReports_Write(report, arguments.json_path, arguments.markdown_path)
    print(
        f"审计完成：{arguments.json_path}；{arguments.markdown_path}",
        file=sys.stdout,
    )
    if not report["release_gate"]["passed"] and not arguments.allow_empty_columns:
        print(
            "发布门禁未通过：至少一个目标字段在有适用公司的市场中整列无数值",
            file=sys.stderr,
        )
        return 3
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
